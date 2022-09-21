import os.path
import shutil
import tkinter as tk
import openpyxl
import openpyxl.styles
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Font, PatternFill
# from openpyxl.styles.borders import Border, Side
from barcode import Code128  # Package called python-barcode
from barcode.writer import ImageWriter  # Requires Pillow
from tkinter import *
from tkinter import filedialog
from tkinterdnd2 import DND_FILES, TkinterDnD

version = '0.1'
file_types = ('xlsx', 'xlsm', 'xltx', 'xltm')


class ChecklistProgram:
    def __init__(self):
        # Initialize main program
        self.root = TkinterDnD.Tk()
        self.root.drop_target_register(DND_FILES)
        self.root.dnd_bind('<<Drop>>', self.open_file)

        # Root window properties
        self.root.geometry('420x420')
        self.root.title('Checklist Generator')
        self.root.resizable(False, False)

        # Variables
        self.cur_dir = os.getcwd() + '/'
        self.batch = None
        self.parts = {}
        self.parts_list = []

        # Button
        self.open_file_button = tk.Button(self.root, text='Browse...', width=10, command=self.open_file, takefocus=0)
        self.open_file_button.pack()

        # Version Label
        self.version_label = Label(self.root, text='ver ' + version)
        self.version_label.place(relx=1, rely=1.01, anchor='se')

        # Main Loop
        self.root.mainloop()

    def open_file(self, event=None):
        if event is None:
            file_path = filedialog.askopenfilename(filetypes=[('Excel File', file_types)])
            if not file_path:
                return
            self.process_file(file_path)
            return
        file_path = event.data.replace('{', '').replace('}', '')
        valid_file = False
        for file_type in file_types:
            if '.' + file_type in file_path:
                valid_file = True
                break
        if not valid_file:
            return
        self.process_file(file_path)

    def process_file(self, file_path: str):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        part_col, type_col, material_col, break_col, weld_col, powder_col, quantity_col, batch_col, setup_col = \
            None, None, None, None, None, None, None, None, None
        read_col_labels = False
        labels_row = None
        no_break_parts = []
        for r in range(1, sheet.max_row + 1):
            current_part = None
            for c in range(1, sheet.max_column + 1):
                cell_obj = sheet.cell(row=r, column=c)
                if read_col_labels:
                    if r != labels_row:
                        read_col_labels = False
                    elif (cell_val := cell_obj.value) == 'Part':
                        part_col = c
                    elif not type_col and cell_val == 'Type':
                        type_col = c
                    elif not material_col and cell_val == 'Material':
                        material_col = c
                    elif not break_col and cell_val == 'Break':
                        break_col = c
                    elif not weld_col and cell_val == 'Weld':
                        weld_col = c
                    elif not powder_col and cell_val == 'PowdCoat Y/N':
                        powder_col = c
                    elif not quantity_col and cell_val == 'Quantity':
                        quantity_col = c
                    elif not batch_col and cell_val == 'Batch':
                        batch_col = c
                    elif not setup_col and cell_val == 'Setup':
                        setup_col = c
                    if read_col_labels:
                        continue
                if part_col and c == part_col and cell_obj.value:
                    current_part = cell_obj.value
                    self.parts_list.append(current_part)
                    self.parts[current_part] = {}
                elif current_part:
                    if type_col and c == type_col:
                        self.parts[current_part]['type'] = cell_obj.value
                    elif material_col and c == material_col:
                        self.parts[current_part]['material'] = cell_obj.value
                    elif break_col and c == break_col:
                        if cell_obj.value == 'N':
                            self.parts[current_part]['break'] = False
                        else:
                            self.parts[current_part]['break'] = True
                    elif weld_col and c == weld_col:
                        if cell_obj.value == 'N':
                            self.parts[current_part]['weld'] = False
                        else:
                            self.parts[current_part]['weld'] = True
                    elif powder_col and c == powder_col:
                        if cell_obj.value == 'N':
                            self.parts[current_part]['powder'] = False
                        else:
                            self.parts[current_part]['powder'] = True
                    elif quantity_col and c == quantity_col:
                        self.parts[current_part]['quantity'] = cell_obj.value
                    elif batch_col and c == batch_col and self.batch is None:
                        self.batch = cell_obj.value
                    elif setup_col and c == setup_col:
                        self.parts[current_part]['setup'] = cell_obj.value
                if cell_obj.value == '#' or cell_obj.value == 'Part':
                    if cell_obj.value == 'Part':
                        part_col = c
                    read_col_labels = True
                    labels_row = r
        if setup_col:
            part_materials, temp_dict, check_list, setup_list, missed_parts = [], {}, [], [], []
            for part in reversed(self.parts_list):
                if not self.parts[part]['break']:
                    no_break_parts.append(part)
                    self.parts_list.pop(self.parts_list.index(part))
                if part not in no_break_parts and self.parts[part]['material'] and \
                        (material := self.parts[part]['material']) not in part_materials:
                    part_materials.append(material)
            part_materials.sort()  # Sort by material
            for material in part_materials:
                temp_dict[material] = []
                for part in self.parts_list:
                    if self.parts[part]['material'] and self.parts[part]['material'] == material:
                        temp_dict[material].append(part)
                        check_list.append(part)
            # Check for missed parts
            for part in self.parts_list:
                if part not in check_list and part not in no_break_parts:
                    missed_parts.append(part)
            parts_list_bkp = [x for x in self.parts_list]
            self.parts_list = []
            for material in part_materials:
                for part in temp_dict[material]:
                    if self.parts[part]['setup'] and (part_setup := self.parts[part]['setup']) not in setup_list:
                        setup_list.append(part_setup)
                setup_list.sort()  # Sort by setup
                for setup in setup_list:
                    for part in temp_dict[material]:
                        if self.parts[part]['setup'] and self.parts[part]['setup'] == setup:
                            self.parts_list.append(part)
            self.parts_list.extend(missed_parts)
            for part in parts_list_bkp:
                if part not in self.parts_list:
                    self.parts_list.append(part)

        self.parts_list.extend(no_break_parts)

        self.make_break_list()

    def make_break_list(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.worksheets[0]
        sheet.freeze_panes = sheet['A2']
        sheet.merge_cells('A1:C1')
        sheet['A1'].value = 'Batch: ' + str(self.batch)
        sheet['A1'].font = Font(bold=True)
        sheet.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(horizontal='center')
        # sheet['D1'].value = 'Notes'
        # sheet['D1'].font = Font(bold=True)
        # sheet.cell(row=1, column=4).alignment = openpyxl.styles.Alignment(horizontal='center')
        temp_dir = self.cur_dir + 'barcode_temp/'
        if os.path.isdir(temp_dir):
            shutil.rmtree(temp_dir)
        os.mkdir(temp_dir)
        alt_bg_color = 'CDCDCD'
        for i, part in enumerate(self.parts_list, start=2):
            part_name = part + ' ['
            if self.parts[part]['break']:
                part_name += 'B'
            if self.parts[part]['weld']:
                if part_name[-1] != '[':
                    part_name += '-'
                part_name += 'W'
            if self.parts[part]['powder']:
                if part_name[-1] != '[':
                    part_name += '-'
                part_name += 'P'
            if part_name.endswith('['):
                part_name = part_name[0:-2]
            else:
                part_name += ']'
            if self.parts[part]['setup']:
                part_name = ''.join([part_name, ' {', self.parts[part]['setup'], '}'])
            sheet.cell(row=i, column=3).alignment = openpyxl.styles.Alignment(vertical='top')
            sheet.cell(row=i, column=3).value = part_name
            if self.parts[part]['weld']:
                font_color = 'E02222'
            elif self.parts[part]['powder']:
                font_color = '0000FF'
            else:
                font_color = '000000'
            sheet.cell(row=i, column=3).font = Font(color=font_color, bold=True)
            sheet.cell(row=i, column=2).alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center')
            sheet.cell(row=i, column=2).value = '☐'
            with open(temp_dir + part + '.png', "wb") as file:
                if i % 2 != 0:
                    Code128(part, writer=ImageWriter()).write(file, options={'write_text': False})
                else:
                    Code128(part, writer=ImageWriter()).write(file, options={'write_text': False,
                                                                             'background': ''.join([
                                                                                 '#', alt_bg_color])})
            img = openpyxl.drawing.image.Image(temp_dir + part + '.png')
            height = 18
            width = 280
            img.height = height
            img.width = width
            # Silly nonsense below
            row_offset = 0.25
            marker = AnchorMarker(col=0, row=i-1, rowOff=cm_to_EMU((row_offset * 49.77) / 99),
                                  colOff=cm_to_EMU((0.1 * 49.77) / 99))
            size = XDRPositiveSize2D(pixels_to_EMU(width), pixels_to_EMU(height))
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            sheet.add_image(img)
        dims = {}
        # border = Border(left=Side(style='thin'), right=Side(style='thin'),
        #                 top=Side(style='thin'), bottom=Side(style='thin'))
        for row in sheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            sheet.column_dimensions[col].width = value + 2
        for row in range(2, sheet.max_row + 1):
            sheet.row_dimensions[row].height = 20
            if row % 2 == 0:
                for col in ['A', 'B', 'C']:
                    sheet[''.join([col, str(row)])].fill = PatternFill(start_color=alt_bg_color, end_color=alt_bg_color,
                                                                       fill_type='solid')
        #     sheet.cell(row=row, column=4).border = border
        sheet.column_dimensions['A'].width = 40
        sheet.column_dimensions['B'].width = 2
        # sheet.column_dimensions['D'].width = 20
        workbook.save('test.xlsx')
        if os.path.isdir(temp_dir):
            shutil.rmtree(temp_dir)
        print('DONE')


ChecklistProgram()
